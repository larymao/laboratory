from pathlib import Path
import json
from datetime import datetime
import pandas as pd
from loguru import logger
from typing import List
from openpyxl.styles import PatternFill

from odps_fetcher import OdpsFetcher, SourceType
from table_comparator import TableComparator

class ComparisonService:
    @staticmethod
    def get_project_root() -> Path:
        """Get the project root directory"""
        return Path(__file__).parent.parent

    @classmethod
    def load_tables_config(cls) -> List[str]:
        """
        Load tables configuration from config.json
        
        Returns:
            List of table names to compare
        """
        config_path = cls.get_project_root() / "config.json"
        if not config_path.exists():
            logger.error(f"Configuration file not found: {config_path}")
            return []
            
        try:
            with open(config_path, 'r') as f:
                config = json.load(f)
                
            tables = config.get('tables', [])
            if not tables:
                logger.error("No tables configured in config.json")
                return []
                
            logger.info(f"Loaded {len(tables)} tables from configuration")
            return tables
        except Exception as e:
            logger.error(f"Error loading configuration: {str(e)}")
            return []

    @staticmethod
    def initialize_comparator() -> TableComparator:
        """
        Initialize the table comparator with data source fetchers
        
        Returns:
            Configured TableComparator instance
        """
        logger.debug("Initializing data source fetchers")
        old_fetcher = OdpsFetcher(SourceType.OLD)
        new_fetcher = OdpsFetcher(SourceType.NEW)
        return TableComparator(old_fetcher, new_fetcher)

    @staticmethod
    def compare_tables(comparator: TableComparator, tables: List[str]) -> None:
        """
        Compare the specified tables using the comparator
        
        Args:
            comparator: The TableComparator instance
            tables: List of table names to compare
        """
        for table in tables:
            logger.info(f"Comparing table: {table}")
            result = comparator.format_comparison_result(table)
            logger.info("\n" + "="*50 + "\n" + result + "\n" + "="*50)

    @staticmethod
    def format_results_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """
        Format the numeric columns in the results DataFrame
        
        Args:
            df: The input DataFrame
            
        Returns:
            Formatted DataFrame
        """
        df = df.copy()
        df['old_count'] = df['old_count'].apply(lambda x: f"{x:,}" if pd.notnull(x) else "N/A")
        df['new_count'] = df['new_count'].apply(lambda x: f"{x:,}" if pd.notnull(x) else "N/A")
        df['diff_absolute'] = df['diff_absolute'].apply(lambda x: f"{x:+,}" if pd.notnull(x) else "N/A")
        df['diff_percentage'] = df['diff_percentage'].apply(lambda x: f"{x:+.2f}%" if pd.notnull(x) else "N/A")
        return df

    @classmethod
    def export_results_to_excel(cls, comparator: TableComparator) -> None:
        """
        Export comparison results to Excel file
        
        Args:
            comparator: The TableComparator instance with results
        """
        if not comparator.comparison_results:
            logger.warning("No comparison results to export")
            return
            
        # Create and format DataFrame
        df = cls.format_results_dataframe(pd.DataFrame(comparator.comparison_results))
        
        # Prepare output path
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        reports_dir = cls.get_project_root() / "reports"
        reports_dir.mkdir(exist_ok=True)
        output_path = str(reports_dir / f"table_comparison_{timestamp}.xlsx")
        
        # Export to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Comparison Results')
            
            # Auto-adjust columns width and highlight differences
            worksheet = writer.sheets['Comparison Results']
            
            # Set yellow background style
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            
            # Iterate through each row, if diff_absolute is not 0, mark the entire row with light yellow
            for row_idx, row in df.iterrows():
                diff_value = row['diff_absolute']
                if diff_value != 'N/A' and diff_value != '0' and diff_value != '+0':
                    # Excel row numbers start from 2 (1 is the header)
                    excel_row = row_idx + 2
                    for col_idx in range(len(df.columns)):
                        cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                        cell.fill = yellow_fill
            
            # Adjust column widths
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).apply(len).max(),
                               len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
        
        logger.info(f"All comparison results exported to: {output_path}")
